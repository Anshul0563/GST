from __future__ import annotations

from copy import copy

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.workbook import Workbook


BLUE = "10244D"
ACCENT = "1746A2"
HEADER = "E8EEF7"
SUBHEADER = "F3F6FA"
ALT_ROW = "FAFBFD"
TOTAL = "FFF4D6"
WHITE = "FFFFFF"
BORDER = "CBD5E1"
TEXT = "0F172A"

thin_side = Side(style="thin", color=BORDER)
medium_side = Side(style="medium", color="94A3B8")
table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def _style(name: str, font: Font, fill: PatternFill, alignment: Alignment, border: Border = table_border, number_format: str = "General") -> NamedStyle:
    style = NamedStyle(name=name)
    style.font = font
    style.fill = fill
    style.alignment = alignment
    style.border = border
    style.number_format = number_format
    return style


def register_styles(workbook: Workbook) -> None:
    existing = {style.name for style in workbook.named_styles if hasattr(style, "name")}
    styles = [
        _style(
            "gst_title",
            Font(name="Calibri", size=13, bold=True, color=WHITE),
            PatternFill("solid", fgColor=BLUE),
            Alignment(horizontal="center", vertical="center"),
            Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side),
        ),
        _style(
            "gst_summary_label",
            Font(name="Calibri", size=10, bold=True, color=TEXT),
            PatternFill("solid", fgColor=SUBHEADER),
            Alignment(horizontal="center", vertical="center", wrap_text=True),
        ),
        _style(
            "gst_summary_value",
            Font(name="Calibri", size=10, bold=True, color=TEXT),
            PatternFill("solid", fgColor=TOTAL),
            Alignment(horizontal="center", vertical="center"),
            number_format="#,##0.00",
        ),
        _style(
            "gst_header",
            Font(name="Calibri", size=10, bold=True, color=TEXT),
            PatternFill("solid", fgColor=HEADER),
            Alignment(horizontal="center", vertical="center", wrap_text=True),
            Border(left=thin_side, right=thin_side, top=medium_side, bottom=medium_side),
        ),
        _style(
            "gst_text",
            Font(name="Calibri", size=10, color=TEXT),
            PatternFill("solid", fgColor=WHITE),
            Alignment(horizontal="left", vertical="center", wrap_text=False),
        ),
        _style(
            "gst_text_alt",
            Font(name="Calibri", size=10, color=TEXT),
            PatternFill("solid", fgColor=ALT_ROW),
            Alignment(horizontal="left", vertical="center", wrap_text=False),
        ),
        _style(
            "gst_number",
            Font(name="Calibri", size=10, color=TEXT),
            PatternFill("solid", fgColor=WHITE),
            Alignment(horizontal="right", vertical="center"),
            number_format="#,##0.00",
        ),
        _style(
            "gst_number_alt",
            Font(name="Calibri", size=10, color=TEXT),
            PatternFill("solid", fgColor=ALT_ROW),
            Alignment(horizontal="right", vertical="center"),
            number_format="#,##0.00",
        ),
        _style(
            "gst_integer",
            Font(name="Calibri", size=10, color=TEXT),
            PatternFill("solid", fgColor=WHITE),
            Alignment(horizontal="right", vertical="center"),
            number_format="0",
        ),
        _style(
            "gst_help",
            Font(name="Calibri", size=9, bold=True, color=WHITE),
            PatternFill("solid", fgColor=ACCENT),
            Alignment(horizontal="center", vertical="center"),
        ),
    ]
    for style in styles:
        if style.name not in existing:
            workbook.add_named_style(style)


def clone_fill(fill: PatternFill) -> PatternFill:
    return copy(fill)
