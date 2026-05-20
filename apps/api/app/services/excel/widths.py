from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


MIN_WIDTH = 10
MAX_WIDTH = 42

HEADER_PRESETS = {
    "GSTIN/UIN of Recipient": 22,
    "GSTIN of E-Commerce Operator": 24,
    "E-Commerce GSTIN": 22,
    "E-Commerce Operator Name": 24,
    "Receiver Name": 24,
    "Invoice Number": 22,
    "Note Number": 22,
    "Sr. No. From": 24,
    "Sr. No. To": 24,
    "Invoice date": 14,
    "Note Date": 14,
    "Place Of Supply": 34,
    "Applicable % of Tax Rate": 18,
    "Rate": 9,
    "Taxable Value": 16,
    "Cess Amount": 14,
    "Invoice Value": 16,
    "Note Value": 16,
    "Total Number": 14,
    "Cancelled": 12,
    "Nature of Supply": 34,
    "Nature of Document": 30,
    "Description": 42,
    "HSN": 14,
    "UQC": 10,
}


def rendered_length(value: object) -> int:
    if value is None:
        return 0
    text = str(value)
    return max((len(part) for part in text.splitlines()), default=0)


def apply_column_widths(ws: Worksheet, header_row: int, max_row: int, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        header = ws.cell(header_row, col_idx).value
        preset = HEADER_PRESETS.get(str(header or ""))
        max_len = preset or MIN_WIDTH
        for row_idx in range(1, max_row + 1):
            max_len = max(max_len, rendered_length(ws.cell(row_idx, col_idx).value) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(MAX_WIDTH, max(MIN_WIDTH, max_len))
