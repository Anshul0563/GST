from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel.sections import SHEET_ORDER, SheetSection, build_sections
from app.services.excel.styles import register_styles, table_border
from app.services.excel.widths import apply_column_widths


NUMERIC_HEADERS = {
    "Total Invoice Value",
    "Total Taxable Value",
    "Total Cess",
    "Invoice Value",
    "Taxable Value",
    "Cess Amount",
    "Note Value",
    "Total Value",
    "Integrated Tax Amount",
    "Central Tax Amount",
    "State/UT Tax Amount",
    "Net value of supplies",
    "Integrated tax",
    "Central tax",
    "State/UT tax",
    "Cess",
}
INTEGER_HEADERS = {"No. of Recipients", "No. of Invoices", "No. of Notes", "No. of HSN", "Total Number", "Cancelled", "Rate", "Total Quantity"}


def write_gstr1_workbook(path: Path, payload: dict[str, Any]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    register_styles(workbook)
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sections = build_sections(payload)
    for sheet_name in SHEET_ORDER:
        section = sections[sheet_name]
        ws = workbook.create_sheet(sheet_name)
        render_section(ws, section)

    workbook.save(path)
    return path


def render_section(ws: Worksheet, section: SheetSection) -> None:
    rows = section.rows
    max_col = max(len(row) for row in rows)
    max_row = len(rows)

    for row_idx, row in enumerate(rows, start=1):
        for col_idx in range(1, max_col + 1):
            value = row[col_idx - 1] if col_idx <= len(row) else None
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = table_border
            apply_cell_style(ws, row_idx, col_idx, section.header_row, max_col)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col - 1)
    ws.cell(1, 1).style = "gst_title"
    ws.cell(1, max_col).style = "gst_help"

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 31
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[section.header_row].height = 34

    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A{section.header_row}:{get_column_letter(max_col)}{max(max_row, section.header_row)}"
    apply_column_widths(ws, section.header_row, max_row, max_col)
    configure_print_layout(ws, max_row, max_col)


def apply_cell_style(ws: Worksheet, row_idx: int, col_idx: int, header_row: int, max_col: int) -> None:
    cell = ws.cell(row_idx, col_idx)
    header_value = ws.cell(header_row, col_idx).value
    if row_idx == 1:
        cell.style = "gst_title" if col_idx < max_col else "gst_help"
        return
    if row_idx in {2, 3}:
        cell.style = "gst_summary_value" if isinstance(cell.value, (int, float)) else "gst_summary_label"
        return
    if row_idx == header_row:
        cell.style = "gst_header"
        return
    if isinstance(cell.value, (int, float)):
        cell.style = "gst_integer" if str(header_value or "") in INTEGER_HEADERS else "gst_number"
    else:
        cell.style = "gst_text_alt" if row_idx % 2 == 1 else "gst_text"
    if col_idx in {1, 2, 3}:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def configure_print_layout(ws: Worksheet, max_row: int, max_col: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "$1:$4"
    ws.print_area = f"A1:{get_column_letter(max_col)}{max_row}"
