from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any
import re

import pandas as pd
from openpyxl import load_workbook

from app.utils.states import STATE_CODES, state_code_from_text


ETIN_MASTER = {
    "amazon": "07AAICA3918J1CV",
    "flipkart": "07AACCF0683K1CU",
    "meesho": "07AARCM9332R1CQ",
}


STATE_ALIASES = {
    "DAMAN": "25",
    "PONDICHERRY": "34",
    "ORISSA": "21",
}


@dataclass
class CalcRecord:
    platform: str
    source: str
    row: int
    gstin: str | None
    etin: str | None
    doc_type: str
    invoice_no: str | None
    order_id: str | None
    order_item_id: str | None
    pos: str | None
    state: str | None
    rate: Decimal
    gross: Decimal
    taxable: Decimal
    igst: Decimal
    cgst: Decimal
    sgst: Decimal
    cess: Decimal = Decimal("0")
    tcs: Decimal | None = None
    tds: Decimal | None = None
    include_in_b2cs: bool = True
    include_in_supeco: bool = True
    include_in_doc_issue: bool = True
    reason: str | None = None

    @property
    def gst(self) -> Decimal:
        return self.igst + self.cgst + self.sgst + self.cess

    @property
    def supply_type(self) -> str:
        return "INTRA" if self.gstin and self.pos and self.gstin[:2] == self.pos else "INTER"


@dataclass
class CalculationResult:
    records: list[CalcRecord] = field(default_factory=list)
    ignored_rows: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    uncertainties: list[str] = field(default_factory=list)
    tax_mismatches: list[dict[str, Any]] = field(default_factory=list)
    missing_fields: list[dict[str, Any]] = field(default_factory=list)
    duplicate_docs: list[dict[str, Any]] = field(default_factory=list)
    flipkart_sheets: list[dict[str, Any]] = field(default_factory=list)


def precise_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    try:
        if pd.isna(value):
            return Decimal("0")
    except TypeError:
        pass
    text = str(value).strip().replace(",", "").replace("₹", "").replace("%", "")
    if text.lower() in {"", "nan", "none", "null", "na"}:
        return Decimal("0")
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return Decimal("0")
    return -amount if negative else amount


def money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def text(value: Any) -> str | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    cleaned = str(value).strip().strip('"')
    return cleaned or None


def pos_code(value: Any) -> str | None:
    raw = text(value)
    if not raw:
        return None
    upper = raw.upper()
    if upper in STATE_ALIASES:
        return STATE_ALIASES[upper]
    if upper.startswith("IN-"):
        upper = upper[3:]
        upper = {
            "DL": "DELHI",
            "MH": "MAHARASHTRA",
            "UP": "UTTAR PRADESH",
            "TN": "TAMIL NADU",
            "KA": "KARNATAKA",
            "GJ": "GUJARAT",
            "RJ": "RAJASTHAN",
            "HR": "HARYANA",
            "WB": "WEST BENGAL",
            "TS": "TELANGANA",
            "AP": "ANDHRA PRADESH",
            "KL": "KERALA",
            "BR": "BIHAR",
            "OR": "ODISHA",
            "OD": "ODISHA",
            "PB": "PUNJAB",
        }.get(upper, upper)
    return state_code_from_text(upper)


def split_total_tax(gstin: str | None, pos: str | None, total_tax: Decimal) -> tuple[Decimal, Decimal, Decimal]:
    if gstin and pos and gstin[:2] == pos:
        half = money(total_tax / Decimal("2"))
        return Decimal("0"), half, total_tax - half
    return total_tax, Decimal("0"), Decimal("0")


def signed_amount(value: Any, doc_type: str) -> Decimal:
    amount = precise_decimal(value)
    if doc_type == "credit_note":
        return -abs(amount)
    if doc_type == "debit_note":
        return abs(amount)
    return amount


def validate_record(result: CalculationResult, record: CalcRecord) -> None:
    if not record.pos or not record.invoice_no or not record.etin:
        result.missing_fields.append({
            "platform": record.platform,
            "source": record.source,
            "row": record.row,
            "invoice_no": record.invoice_no,
            "state": record.state,
            "pos": record.pos,
            "etin": record.etin,
        })
    expected = money(record.taxable * record.rate / Decimal("100"))
    actual = money(record.gst)
    if abs(expected - actual) > Decimal("1.00"):
        result.tax_mismatches.append({
            "platform": record.platform,
            "source": record.source,
            "row": record.row,
            "invoice_no": record.invoice_no,
            "expected": float(expected),
            "actual": float(actual),
            "taxable": float(money(record.taxable)),
            "rate": float(record.rate),
        })


def parse_meesho(result: CalculationResult, sales_path: Path, returns_path: Path, invoice_path: Path) -> None:
    invoice_df = pd.read_excel(invoice_path, dtype=object)
    invoice_map = {
        text(row["Suborder No."]): text(row["Invoice No."])
        for _, row in invoice_df.iterrows()
    }
    for doc_type, path in (("invoice", sales_path), ("credit_note", returns_path)):
        df = pd.read_excel(path, dtype=object)
        for index, row in df.iterrows():
            gstin = text(row.get("gstin"))
            pos = pos_code(row.get("end_customer_state_new"))
            taxable = signed_amount(row.get("total_taxable_sale_value"), doc_type)
            total_tax = signed_amount(row.get("tax_amount"), doc_type)
            gross = signed_amount(row.get("total_invoice_value"), doc_type)
            igst, cgst, sgst = split_total_tax(gstin, pos, total_tax)
            record = CalcRecord(
                platform="Meesho",
                source=path.name,
                row=index + 2,
                gstin=gstin,
                etin=ETIN_MASTER["meesho"],
                doc_type=doc_type,
                invoice_no=invoice_map.get(text(row.get("sub_order_num"))) or text(row.get("sub_order_num")),
                order_id=text(row.get("sub_order_num")),
                order_item_id=text(row.get("sub_order_num")),
                pos=pos,
                state=text(row.get("end_customer_state_new")),
                rate=precise_decimal(row.get("gst_rate")),
                gross=gross,
                taxable=taxable,
                igst=igst,
                cgst=cgst,
                sgst=sgst,
                tcs=None,
                tds=None,
            )
            result.records.append(record)
            validate_record(result, record)
    result.warnings.append("Meesho: source has total GST only; IGST/CGST/SGST is split after summing by supply type/POS.")


def parse_amazon(result: CalculationResult, path: Path) -> None:
    df = pd.read_csv(path, dtype=object)
    for index, row in df.iterrows():
        transaction_type = text(row.get("Transaction Type"))
        gross = precise_decimal(row.get("Invoice Amount"))
        taxable = precise_decimal(row.get("Tax Exclusive Gross"))
        total_tax = precise_decimal(row.get("Total Tax Amount"))
        if transaction_type == "Cancel" and gross == 0 and taxable == 0 and total_tax == 0:
            result.ignored_rows.append({
                "platform": "Amazon",
                "source": path.name,
                "row": index + 2,
                "reason": "Cancel row with zero amount/no invoice number",
                "order_id": text(row.get("Order Id")),
            })
            continue
        gstin = text(row.get("Seller Gstin"))
        igst = precise_decimal(row.get("Igst Tax")) + precise_decimal(row.get("Shipping Igst Tax")) + precise_decimal(row.get("Gift Wrap Igst Tax"))
        cgst = precise_decimal(row.get("Cgst Tax")) + precise_decimal(row.get("Shipping Cgst Tax")) + precise_decimal(row.get("Gift Wrap Cgst Tax"))
        sgst = precise_decimal(row.get("Sgst Tax")) + precise_decimal(row.get("Shipping Sgst Tax")) + precise_decimal(row.get("Utgst Tax")) + precise_decimal(row.get("Shipping Utgst Tax"))
        igst_rate = precise_decimal(row.get("Igst Rate"))
        rate = igst_rate * Decimal("100") if Decimal("0") < igst_rate < Decimal("1") else igst_rate
        record = CalcRecord(
            platform="Amazon",
            source=path.name,
            row=index + 2,
            gstin=gstin,
            etin=ETIN_MASTER["amazon"],
            doc_type="invoice",
            invoice_no=text(row.get("Invoice Number")),
            order_id=text(row.get("Order Id")),
            order_item_id=text(row.get("Shipment Item Id")),
            pos=pos_code(row.get("Ship To State")),
            state=text(row.get("Ship To State")),
            rate=rate,
            gross=gross,
            taxable=taxable,
            igst=igst,
            cgst=cgst,
            sgst=sgst,
            cess=precise_decimal(row.get("Compensatory Cess Tax")),
            tcs=precise_decimal(row.get("Tcs Cgst Amount")) + precise_decimal(row.get("Tcs Sgst Amount")) + precise_decimal(row.get("Tcs Utgst Amount")) + precise_decimal(row.get("Tcs Igst Amount")),
            tds=None,
        )
        result.records.append(record)
        validate_record(result, record)


def parse_flipkart(result: CalculationResult, path: Path) -> None:
    workbook = load_workbook(path, read_only=False, data_only=True)
    result.flipkart_sheets = [
        {"title": sheet.title, "state": sheet.sheet_state, "max_row": sheet.max_row, "max_col": sheet.max_column}
        for sheet in workbook.worksheets
    ]
    sales_df = pd.read_excel(path, sheet_name="Sales Report", dtype=object)
    for index, row in sales_df.iterrows():
        event_type = text(row.get("Event Type"))
        if not event_type:
            result.ignored_rows.append({"platform": "Flipkart", "source": path.name, "row": index + 2, "reason": "Blank event type"})
            continue
        doc_type = "credit_note" if event_type.lower() == "return" else "invoice"
        taxable = signed_amount(row.get("Taxable Value (Final Invoice Amount -Taxes)"), doc_type)
        gross = signed_amount(row.get("Final Invoice Amount (Price after discount+Shipping Charges)"), doc_type)
        record = CalcRecord(
            platform="Flipkart",
            source=f"{path.name}:Sales Report",
            row=index + 2,
            gstin=text(row.get("Seller GSTIN")),
            etin=ETIN_MASTER["flipkart"],
            doc_type=doc_type,
            invoice_no=text(row.get("Buyer Invoice ID")),
            order_id=text(row.get("Order ID")),
            order_item_id=text(row.get("Order Item ID")),
            pos=pos_code(row.get("Customer's Delivery State")) or pos_code(row.get("Customer's Billing State")),
            state=text(row.get("Customer's Delivery State")),
            rate=precise_decimal(row.get("IGST Rate")) or precise_decimal(row.get("CGST Rate")) + precise_decimal(row.get("SGST Rate (or UTGST as applicable)")),
            gross=gross,
            taxable=taxable,
            igst=precise_decimal(row.get("IGST Amount")),
            cgst=precise_decimal(row.get("CGST Amount")),
            sgst=precise_decimal(row.get("SGST Amount (Or UTGST as applicable)")),
            cess=precise_decimal(row.get("Luxury Cess Amount")),
            tcs=precise_decimal(row.get("Total TCS Deducted")),
            tds=precise_decimal(row.get("TDS Amount")),
        )
        result.records.append(record)
        validate_record(result, record)

    cashback_df = pd.read_excel(path, sheet_name="Cash Back Report", dtype=object)
    for index, row in cashback_df.iterrows():
        doc_no = text(row.get("Credit Note ID/ Debit Note ID"))
        if not doc_no:
            result.ignored_rows.append({"platform": "Flipkart", "source": path.name, "row": index + 2, "reason": "Cashback row without document number"})
            continue
        doc_type = "debit_note" if doc_no.startswith("LZAA") else "credit_note" if doc_no.startswith("LYAA") else "credit_note"
        # Flipkart cashback report already carries the sign used by the GST tool.
        # LYAA rows are positive and LZAA rows are negative in this source.
        taxable = precise_decimal(row.get("Taxable Value"))
        gross = precise_decimal(row.get("Invoice Amount"))
        record = CalcRecord(
            platform="Flipkart",
            source=f"{path.name}:Cash Back Report",
            row=index + 2,
            gstin=text(row.get("Seller GSTIN")),
            etin=ETIN_MASTER["flipkart"],
            doc_type=doc_type,
            invoice_no=doc_no,
            order_id=text(row.get("Order ID")),
            order_item_id=text(row.get("Order Item ID")),
            pos=pos_code(row.get("Customer's Delivery State")),
            state=text(row.get("Customer's Delivery State")),
            rate=precise_decimal(row.get("IGST Rate")) or precise_decimal(row.get("CGST Rate")) + precise_decimal(row.get("SGST Rate (or UTGST as applicable)")),
            gross=gross,
            taxable=taxable,
            igst=precise_decimal(row.get("IGST Amount")),
            cgst=precise_decimal(row.get("CGST Amount")),
            sgst=precise_decimal(row.get("SGST Amount (Or UTGST as applicable)")),
            cess=precise_decimal(row.get("Luxury Cess Amount")),
            tcs=precise_decimal(row.get("Total TCS Deducted")),
            tds=precise_decimal(row.get("TDS Amount")),
            include_in_b2cs=True,
            include_in_supeco=True,
            include_in_doc_issue=True,
            reason="Flipkart cashback GST adjustment included with source sign.",
        )
        result.records.append(record)
        validate_record(result, record)


def aggregate(records: list[CalcRecord]) -> dict[str, Any]:
    invoices = [r for r in records if r.doc_type == "invoice"]
    credits = [r for r in records if r.doc_type == "credit_note"]
    debits = [r for r in records if r.doc_type == "debit_note"]
    return {
        "rows": len(records),
        "invoices": len(invoices),
        "credit_notes": len(credits),
        "debit_notes": len(debits),
        "gross_sales": money(sum((r.gross for r in invoices), Decimal("0"))),
        "sales_taxable_before_returns": money(sum((r.taxable for r in invoices), Decimal("0"))),
        "return_taxable": money(sum((r.taxable for r in credits), Decimal("0"))),
        "debit_taxable": money(sum((r.taxable for r in debits), Decimal("0"))),
        "taxable": money(sum((r.taxable for r in records), Decimal("0"))),
        "igst": money(sum((r.igst for r in records), Decimal("0"))),
        "cgst": money(sum((r.cgst for r in records), Decimal("0"))),
        "sgst": money(sum((r.sgst for r in records), Decimal("0"))),
        "cess": money(sum((r.cess for r in records), Decimal("0"))),
        "total_gst": money(sum((r.gst for r in records), Decimal("0"))),
        "invoice_value": money(sum((r.gross for r in records), Decimal("0"))),
        "tcs": None if all(r.tcs is None for r in records) else money(sum((r.tcs or Decimal("0") for r in records), Decimal("0"))),
        "tds": None if all(r.tds is None for r in records) else money(sum((r.tds or Decimal("0") for r in records), Decimal("0"))),
    }


def grouped(records: list[CalcRecord], key_fn) -> dict[str, dict[str, Any]]:
    groups: dict[str, list[CalcRecord]] = defaultdict(list)
    for record in records:
        groups[key_fn(record)].append(record)
    return {key: aggregate(value) for key, value in sorted(groups.items())}


def document_ranges(records: list[CalcRecord]) -> dict[str, list[dict[str, Any]]]:
    output: dict[str, list[dict[str, Any]]] = {"invoice": [], "credit_note": [], "debit_note": []}

    def inclusive_series(docs: list[str], pattern: str, prefix_group: int = 1, number_group: int = 2) -> dict[str, Any] | None:
        matches = []
        for doc in docs:
            match = re.match(pattern, doc)
            if match:
                matches.append((match.group(prefix_group), int(match.group(number_group)), doc))
        if not matches:
            return None
        prefix = matches[0][0]
        start = min(item[1] for item in matches)
        end = max(item[1] for item in matches)
        return {"from": f"{prefix}{start}", "to": f"{prefix}{end}", "count": end - start + 1}

    def actual_series(docs: list[str]) -> dict[str, Any] | None:
        unique_docs = sorted(set(docs), key=lambda item: (len(item), item))
        if not unique_docs:
            return None
        return {"from": unique_docs[0], "to": unique_docs[-1], "count": len(unique_docs)}

    meesho_invoices = [r.invoice_no for r in records if r.platform == "Meesho" and r.doc_type == "invoice" and r.invoice_no]
    meesho_credits = [r.invoice_no for r in records if r.platform == "Meesho" and r.doc_type == "credit_note" and r.invoice_no]
    meesho_invoice_range = inclusive_series(meesho_invoices, r"^(6p5kc27)(\d+)$")
    meesho_credit_range = inclusive_series(meesho_credits, r"^(6p5kc27C)(\d+)$")
    if meesho_invoice_range:
        output["invoice"].append(meesho_invoice_range)
    if meesho_credit_range:
        output["credit_note"].append(meesho_credit_range)

    amazon_invoices = [r.invoice_no for r in records if r.platform == "Amazon" and r.doc_type == "invoice" and r.invoice_no]
    amazon_range = actual_series(amazon_invoices)
    if amazon_range:
        output["invoice"].append(amazon_range)

    flip_invoices = [r.invoice_no for r in records if r.platform == "Flipkart" and r.doc_type == "invoice" and r.invoice_no and r.invoice_no.startswith("LWAB")]
    flip_returns = [r.invoice_no for r in records if r.platform == "Flipkart" and r.doc_type == "credit_note" and r.invoice_no and r.invoice_no.startswith("MFAB")]
    flip_cashback_credits = [r.invoice_no for r in records if r.platform == "Flipkart" and r.doc_type == "credit_note" and r.invoice_no and r.invoice_no.startswith("LYAA")]
    flip_debits = [r.invoice_no for r in records if r.platform == "Flipkart" and r.doc_type == "debit_note" and r.invoice_no and r.invoice_no.startswith("LZAA")]
    for docs, doc_type in (
        (flip_invoices, "invoice"),
        (flip_returns, "credit_note"),
        (flip_cashback_credits, "credit_note"),
        (flip_debits, "debit_note"),
    ):
        item = actual_series(docs)
        if item:
            output[doc_type].append(item)

    for ranges in output.values():
        ranges.sort(key=lambda item: item["from"])
    return output


def calculate_marketplace_summary(paths: dict[str, str | Path]) -> dict[str, Any]:
    result = CalculationResult()
    parse_meesho(result, Path(paths["meesho_sales"]), Path(paths["meesho_returns"]), Path(paths["meesho_invoice"]))
    parse_amazon(result, Path(paths["amazon"]))
    parse_flipkart(result, Path(paths["flipkart"]))

    key_counter = Counter((r.platform, r.invoice_no, r.order_item_id, r.doc_type) for r in result.records if r.invoice_no or r.order_item_id)
    result.duplicate_docs = [
        {"platform": key[0], "invoice_no": key[1], "order_item_id": key[2], "doc_type": key[3], "count": count}
        for key, count in key_counter.items()
        if count > 1
    ]

    included = [record for record in result.records if record.include_in_b2cs]
    platform_summary = grouped(included, lambda record: record.platform)
    b2cs_summary = grouped(included, lambda record: f"{record.supply_type}|rate={money(record.rate)}|pos={record.pos}")
    supeco_summary = grouped([record for record in result.records if record.include_in_supeco], lambda record: f"{record.platform}|{record.etin or 'UNKNOWN'}")
    doc_ranges = document_ranges(result.records)

    return {
        "combined": aggregate(included),
        "platform_summary": platform_summary,
        "b2cs_summary": b2cs_summary,
        "supeco_summary": supeco_summary,
        "document_ranges": doc_ranges,
        "document_counts": {
            "invoice": sum(item["count"] for item in doc_ranges.get("invoice", [])),
            "credit_note": sum(item["count"] for item in doc_ranges.get("credit_note", [])),
            "debit_note": sum(item["count"] for item in doc_ranges.get("debit_note", [])),
        },
        "ignored_rows": result.ignored_rows,
        "missing_fields": result.missing_fields,
        "tax_mismatches": result.tax_mismatches,
        "duplicate_docs": result.duplicate_docs,
        "warnings": result.warnings,
        "uncertainties": result.uncertainties,
        "flipkart_sheets": result.flipkart_sheets,
        "records": result.records,
    }


def serializable(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(money(value))
    if isinstance(value, CalcRecord):
        return serializable(value.__dict__)
    if isinstance(value, dict):
        return {str(key): serializable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [serializable(item) for item in value]
    return value
