from __future__ import annotations

from copy import copy
from pathlib import Path
import os
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel.sections import SHEET_ORDER, build_sections


DEFAULT_TEMPLATE_PATH = Path("/home/jarvis/Downloads/GSTR1_07TCRPS8655B1ZK_monthly_032026.xlsx")


def template_path() -> Path | None:
    configured = os.getenv("GSTTOOL_EXCEL_TEMPLATE_PATH")
    if configured:
        path = Path(configured)
        return path if path.exists() else None
    return DEFAULT_TEMPLATE_PATH if DEFAULT_TEMPLATE_PATH.exists() else None


def write_gstr1_template_excel(
    path: Path,
    payload: dict[str, Any],
    source_template: Path | None = None,
) -> Path:
    source = source_template or template_path()
    if source is None:
        raise FileNotFoundError("GSTTool Excel template workbook is not available")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source)
    sections = build_sections(payload)

    for sheet_name in SHEET_ORDER:
        if sheet_name not in workbook.sheetnames:
            continue
        sync_sheet_values(workbook[sheet_name], sections[sheet_name].rows)

    workbook.save(path)
    return path


def sync_sheet_values(ws: Worksheet, rows: list[list[Any]]) -> None:
    target_rows = len(rows)
    target_cols = max(max((len(row) for row in rows), default=1), ws.max_column)
    fit_row_count(ws, target_rows, target_cols)

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, target_cols + 1):
            ws.cell(row_idx, col_idx).value = None

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx).value = value


def fit_row_count(ws: Worksheet, target_rows: int, target_cols: int) -> None:
    current_rows = ws.max_row
    if target_rows > current_rows:
        for row_idx in range(current_rows + 1, target_rows + 1):
            ws.insert_rows(row_idx)
            copy_row_layout(ws, min(current_rows, max(5, current_rows)), row_idx, target_cols)
    elif target_rows < current_rows:
        ws.delete_rows(target_rows + 1, current_rows - target_rows)


def copy_row_layout(ws: Worksheet, source_row: int, target_row: int, target_cols: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col_idx in range(1, target_cols + 1):
        source = ws.cell(source_row, col_idx)
        target = ws.cell(target_row, col_idx)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
