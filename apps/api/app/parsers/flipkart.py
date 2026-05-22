from pathlib import Path
from datetime import date

from openpyxl import load_workbook
import pandas as pd

from app.parsers.base import (
    MarketplaceParser,
    ParseResult,
    clean_column,
    detect_header_row_frame,
    first_value,
    has_explicit_tax_split,
    should_skip_transaction,
    unique_headers,
)
from app.services.pos_resolver import (
    new_pos_debug,
    observe_pos_debug,
    resolve_pos,
)
from app.services.transaction_normalizer import finalize_transaction


class FlipkartParser(MarketplaceParser):
    platform = "flipkart"

    def _period_bounds(self) -> tuple[date, date, date]:
        month = int(self.filing_period[:2])
        year = int(self.filing_period[2:])
        start = date(year, month, 1)
        next_month = date(
            year + (1 if month == 12 else 0),
            1 if month == 12 else month + 1,
            1,
        )
        following_month = date(
            next_month.year + (1 if next_month.month == 12 else 0),
            1 if next_month.month == 12 else next_month.month + 1,
            1,
        )
        return start, next_month, following_month

    def _series_key(self, txn: dict, sheet_name: str) -> str:
        invoice_no = str(txn.get("invoice_no") or "").upper()
        sheet = sheet_name.lower()
        if "cash back report" in sheet:
            if invoice_no.startswith("CAN"):
                return "legacy_cashback_credit"
            if invoice_no.startswith("DAL"):
                return "legacy_cashback_debit"
            if invoice_no.startswith("LYAA"):
                return "new_cashback_credit"
            if invoice_no.startswith("LZAA"):
                return "new_cashback_debit"
            return f"cashback:{txn.get('doc_type')}"
        if invoice_no.startswith("FAWRLX"):
            return "legacy_sales"
        if invoice_no.startswith("RAT"):
            return "legacy_returns"
        if invoice_no.startswith("LWAB"):
            return "new_sales"
        if invoice_no.startswith("MFAB"):
            return "new_returns"
        return f"sales:{txn.get('doc_type')}"

    def _report_cycle_filter(
        self, candidates: list[dict]
    ) -> tuple[set[int], dict[int, str]]:
        period_start, next_month, following_month = self._period_bounds()
        included: set[int] = set()
        reasons: dict[int, str] = {}
        current_rows: dict[str, list[tuple[int, date]]] = {}
        next_rows: dict[str, list[tuple[int, date]]] = {}
        has_new_series_in_period = False

        for index, item in enumerate(candidates):
            document_date = item["txn"].get("document_date")
            if not isinstance(document_date, date):
                reasons[index] = "document date unavailable"
                continue
            series = self._series_key(item["txn"], item["sheet_name"])
            if period_start <= document_date < next_month:
                included.add(index)
                reasons[index] = "document date is inside filing period"
                current_rows.setdefault(series, []).append((index, document_date))
                if series.startswith("new_"):
                    has_new_series_in_period = True
            elif next_month <= document_date < following_month:
                next_rows.setdefault(series, []).append((index, document_date))
                reasons[index] = "document date outside filing period"
            else:
                reasons[index] = "document date outside filing period"

        if not has_new_series_in_period:
            for series in (
                "legacy_sales",
                "legacy_returns",
                "legacy_cashback_credit",
            ):
                ordered = sorted(current_rows.get(series, []), key=lambda item: item[1])
                cutoff = None
                previous = None
                for _, document_date in ordered:
                    if previous and (document_date - previous).days >= 5:
                        cutoff = document_date
                        break
                    previous = document_date
                if cutoff is None:
                    continue
                for index, document_date in ordered:
                    if document_date < cutoff:
                        included.discard(index)
                        reasons[index] = "Flipkart report-cycle pre-window row"
                    else:
                        included.add(index)
                        reasons[index] = "Flipkart report-cycle current window row"

            for series in ("legacy_sales", "legacy_cashback_credit"):
                ordered = sorted(next_rows.get(series, []), key=lambda item: item[1])
                if not ordered:
                    continue
                previous = ordered[0][1]
                for index, document_date in ordered:
                    if (document_date - previous).days >= 5:
                        break
                    included.add(index)
                    reasons[index] = "Flipkart report-cycle next-month opening row"
                    previous = document_date

        return included, reasons

    def _classify_doc_type(self, row: dict, sheet_name: str) -> str:
        sheet = sheet_name.lower()
        if "cash back report" in sheet:
            document_type = str(
                first_value(row, ["document type", "doc type", "type"]) or ""
            ).lower()
            if "debit" in document_type:
                return "debit_note"
            return "credit_note"

        event_type = str(first_value(row, ["event type"]) or "").lower()
        if "return" in event_type or "cancellation" in event_type:
            return "credit_note"
        return "invoice"

    def _debug_row(
        self,
        result: ParseResult,
        *,
        file_name: str,
        sheet_name: str,
        row_number: int,
        txn: dict,
        source_row: dict,
        included: bool,
        reason: str,
        running_total: dict,
    ) -> None:
        result.debug.setdefault("row_level_debug", []).append(
            {
                "source_file": file_name,
                "sheet": sheet_name,
                "row": row_number,
                "invoice_no": txn.get("invoice_no"),
                "event_type": first_value(source_row, ["event type"]),
                "document_type": first_value(
                    source_row,
                    ["document type", "doc type", "type"],
                ),
                "doc_type": txn.get("doc_type"),
                "document_date": str(txn.get("document_date")),
                "taxable_value": str(txn.get("taxable_value")),
                "igst": str(txn.get("igst")),
                "cgst": str(txn.get("cgst")),
                "sgst": str(txn.get("sgst")),
                "included": included,
                "reason": reason,
                "running_total": {
                    "taxable_value": str(running_total["taxable_value"]),
                    "igst": str(running_total["igst"]),
                    "cgst": str(running_total["cgst"]),
                    "sgst": str(running_total["sgst"]),
                },
            }
        )

    def parse(self, files: list[Path]) -> ParseResult:
        result = ParseResult()
        result.debug = new_pos_debug(self.platform)
        candidates: list[dict] = []
        running_total = {
            "taxable_value": 0,
            "igst": 0,
            "cgst": 0,
            "sgst": 0,
        }

        for path in files:
            try:
                workbook = load_workbook(
                    path,
                    data_only=True,
                    read_only=False,
                )

                for sheet in workbook.worksheets:
                    rows = list(sheet.iter_rows(values_only=True))

                    if not rows:
                        continue

                    frame = pd.DataFrame(rows)

                    header_index = detect_header_row_frame(
                        frame,
                        ["order", "invoice", "taxable", "igst", "cgst", "sgst"],
                    )

                    result.debug["header_rows"].append(
                        {
                            "file": path.name,
                            "sheet": sheet.title,
                            "header_row": int(header_index) + 1,
                        }
                    )

                    headers = unique_headers(
                        [
                            clean_column(value) or f"column {idx}"
                            for idx, value in enumerate(
                                frame.iloc[header_index].tolist()
                            )
                        ]
                    )

                    data = frame.iloc[header_index + 1 :].copy()
                    data.columns = headers
                    data = data.dropna(how="all")

                    if data.empty:
                        continue

                    for index, series in data.iterrows():
                        row = series.to_dict()

                        txn = self.normalize_row(
                            row,
                            f"{path.name}:{sheet.title}",
                        )
                        txn["_preserve_source_sign"] = True
                        txn["doc_type"] = self._classify_doc_type(row, sheet.title)
                        if has_explicit_tax_split(row):
                            txn["_preserve_source_tax_split"] = True

                        observe_pos_debug(
                            result.debug,
                            int(index) + 1,
                            resolve_pos(row, txn, self.platform),
                            row,
                        )

                        source_row_number = int(index) + 1

                        if should_skip_transaction(txn):
                            self._debug_row(
                                result,
                                file_name=path.name,
                                sheet_name=sheet.title,
                                row_number=source_row_number,
                                txn=txn,
                                source_row=row,
                                included=False,
                                reason="empty transaction row",
                                running_total=running_total,
                            )
                            continue

                        finalized = finalize_transaction(txn)
                        candidates.append(
                            {
                                "txn": finalized,
                                "source_row": row,
                                "file_name": path.name,
                                "sheet_name": sheet.title,
                                "source_row_number": source_row_number,
                            }
                        )

            except Exception as exc:
                result.errors.append(
                    {
                        "file": path.name,
                        "error": str(exc),
                    }
                )

        included, reasons = self._report_cycle_filter(candidates)
        for index, item in enumerate(candidates):
            txn = item["txn"]
            is_included = index in included
            reason = reasons.get(index, "document date outside filing period")
            if is_included:
                result.transactions.append(txn)
                running_total["taxable_value"] += txn.get("taxable_value", 0)
                running_total["igst"] += txn.get("igst", 0)
                running_total["cgst"] += txn.get("cgst", 0)
                running_total["sgst"] += txn.get("sgst", 0)
            else:
                result.debug.setdefault("period_excluded_rows", []).append(
                    {
                        "file": item["file_name"],
                        "sheet": item["sheet_name"],
                        "row": item["source_row_number"],
                        "invoice_no": txn.get("invoice_no"),
                        "doc_type": txn.get("doc_type"),
                        "document_date": str(txn.get("document_date")),
                        "taxable_value": str(txn.get("taxable_value")),
                        "reason": reason,
                    }
                )
            self._debug_row(
                result,
                file_name=item["file_name"],
                sheet_name=item["sheet_name"],
                row_number=item["source_row_number"],
                txn=txn,
                source_row=item["source_row"],
                included=is_included,
                reason=reason,
                running_total=running_total,
            )

        return result
