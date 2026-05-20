import copy
import json
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from app.parsers.amazon import AmazonParser
from app.parsers.flipkart import FlipkartParser
from app.parsers.meesho import MeeshoParser
from app.services.gst import GSTTOOL_COMPATIBLE, build_gstr1_json
from app.services.gsttool_parity_validator import compare_against_reference
from app.services.excel_template_export import write_gstr1_template_excel


class GstToolGoldenJsonTests(unittest.TestCase):
    gstin = "07TCRPS8655B1ZK"

    def golden_cases(self):
        return {
            "022026": {
                "reference": Path("/home/jarvis/Downloads/GSTR1_returns_07TCRPS8655B1ZK_monthly_022026.json"),
                "amazon": [Path("storage/uploads/4/7/eef204f91f474184b31853737577779d.csv")],
                "flipkart": [Path("storage/uploads/4/5/fee8bd723cf04b8bb6637402f6be0c26.xlsx")],
                "meesho": [
                    Path("storage/uploads/4/6/a5ecca73d9d348c592a6f640b5e61542.xlsx"),
                    Path("storage/uploads/4/6/e5d97d93eb254b4d8a73bab35f2b01d0.xlsx"),
                    Path("storage/uploads/4/6/2ce7ac3ccd9c4ca5a2ea5a84a76cd8bf.xlsx"),
                ],
            },
            "032026": {
                "reference": Path("/home/jarvis/Downloads/GSTR1_returns_07TCRPS8655B1ZK_monthly_032026.json"),
                "amazon": [Path("/home/jarvis/Downloads/MTR_B2C-MARCH-2026-A1YGIWFZR88S6S.csv")],
                "flipkart": [Path("/home/jarvis/Downloads/11d11828-0221-4866-b714-b5a26595f116_1779106486000.xlsx")],
                "meesho": [
                    Path("/home/jarvis/Downloads/tcs_sales.xlsx"),
                    Path("/home/jarvis/Downloads/tcs_sales_return.xlsx"),
                    Path("/home/jarvis/Downloads/Tax_invoice_details.xlsx"),
                ],
            },
            "042026": {
                "reference": Path("/home/jarvis/Downloads/GSTR1_returns_07TCRPS8655B1ZK_monthly_042026.json"),
                "amazon": [],
                "flipkart": [],
                "meesho": [],
            },
        }

    def payload_for_period(self, period: str):
        case = self.golden_cases()[period]
        reference_path = case["reference"]
        raw_paths = case["amazon"] + case["flipkart"] + case["meesho"]
        missing = [path for path in [reference_path, *raw_paths] if not path.exists()]
        if missing:
            self.skipTest(f"GSTTool golden fixtures are not available: {missing}")
        if not raw_paths:
            self.skipTest(f"Raw marketplace files for {period} are not available.")
        rows = []
        rows.extend(AmazonParser(self.gstin, period).parse(case["amazon"]).transactions)
        rows.extend(FlipkartParser(self.gstin, period).parse(case["flipkart"]).transactions)
        rows.extend(MeeshoParser(self.gstin, period).parse(case["meesho"]).transactions)
        reference = json.loads(reference_path.read_text(encoding="utf-8"))
        source_documents = {str(row.get("invoice_no")) for row in rows if row.get("invoice_no")}
        reference_endpoints = {
            str(doc[field])
            for section in reference.get("doc_issue", {}).get("doc_det", [])
            for doc in section.get("docs", [])
            for field in ("from", "to")
        }
        missing_reference_documents = sorted(reference_endpoints - source_documents)
        if missing_reference_documents:
            self.skipTest(
                f"Raw marketplace fixture set for {period} does not match original GSTTool JSON; "
                f"missing document endpoints: {missing_reference_documents[:6]}"
            )
        return build_gstr1_json(self.gstin, period, rows, GSTTOOL_COMPATIBLE)

    def assert_exact_gsttool_json(self, period: str):
        reference_path = self.golden_cases()[period]["reference"]
        generated = self.payload_for_period(period)
        reference = json.loads(reference_path.read_text(encoding="utf-8"))
        report = compare_against_reference(reference, generated)

        self.assertEqual(list(generated.keys()), list(reference.keys()))
        self.assertEqual(generated.get("hash"), reference.get("hash"))
        self.assertEqual(generated.get("b2cs"), reference.get("b2cs"))
        self.assertEqual(generated.get("supeco", {}).get("clttx"), reference.get("supeco", {}).get("clttx"))
        self.assertEqual(generated.get("doc_issue"), reference.get("doc_issue"))
        self.assertTrue(report["exact_match"], report["differences"][:20])
        self.assertEqual(report["match_score"], 100.0)

    def march_payload(self):
        gstin = "07TCRPS8655B1ZK"
        period = "032026"
        raw_paths = {
            "amazon": Path("/home/jarvis/Downloads/MTR_B2C-MARCH-2026-A1YGIWFZR88S6S.csv"),
            "flipkart": Path("/home/jarvis/Downloads/11d11828-0221-4866-b714-b5a26595f116_1779106486000.xlsx"),
            "meesho_sales": Path("/home/jarvis/Downloads/tcs_sales.xlsx"),
            "meesho_returns": Path("/home/jarvis/Downloads/tcs_sales_return.xlsx"),
            "meesho_invoice": Path("/home/jarvis/Downloads/Tax_invoice_details.xlsx"),
        }
        if not all(path.exists() for path in raw_paths.values()):
            self.skipTest("March raw marketplace files are not available.")
        rows = []
        rows.extend(AmazonParser(gstin, period).parse([raw_paths["amazon"]]).transactions)
        rows.extend(FlipkartParser(gstin, period).parse([raw_paths["flipkart"]]).transactions)
        rows.extend(
            MeeshoParser(gstin, period)
            .parse([raw_paths["meesho_sales"], raw_paths["meesho_returns"], raw_paths["meesho_invoice"]])
            .transactions
        )
        return build_gstr1_json(gstin, period, rows, GSTTOOL_COMPATIBLE)

    def test_february_raw_marketplace_files_rebuild_original_gsttool_json_exactly(self):
        self.assert_exact_gsttool_json("022026")

    def test_march_raw_marketplace_files_rebuild_original_gsttool_json_exactly(self):
        self.assert_exact_gsttool_json("032026")

    def test_april_raw_marketplace_files_rebuild_original_gsttool_json_exactly(self):
        self.assert_exact_gsttool_json("042026")

    def test_original_march_json_preserves_gsttool_quirks(self):
        path = Path("/home/jarvis/Downloads/GSTR1_returns_07TCRPS8655B1ZK_monthly_032026.json")
        if not path.exists():
            self.skipTest("Original March GSTTool JSON is not available on this machine.")

        payload = json.loads(path.read_text(encoding="utf-8"))
        zero_pos = {row["pos"] for row in payload["b2cs"] if row.get("txval") == 0 and row.get("iamt") == 0}
        doc_ranges = {
            (doc["from"], doc["to"])
            for section in payload["doc_issue"]["doc_det"]
            for doc in section["docs"]
        }
        report = compare_against_reference(payload, copy.deepcopy(payload))

        self.assertTrue({"18", "04", "06", "20"}.issubset(zero_pos))
        self.assertIn(("FAWRLX2600000080", "LWABOG7260000005"), doc_ranges)
        self.assertIn(("MFABNVY260000001", "RAT6SO2600000034"), doc_ranges)
        self.assertIn(("CANQ1W2600000013", "LYAA9U7260000001"), doc_ranges)
        self.assertIn(("DAL84U2600000005", "LZAA9B7260000001"), doc_ranges)
        self.assertTrue(report["exact_match"])
        self.assertEqual(report["match_score"], 100.0)

    def test_april_golden_json_matches_parity_contract(self):
        path = Path("exports/gst_bharat_gstr1_07TCRPS8655B1ZK_042026.json")
        if not path.exists():
            self.skipTest("April GSTTool-compatible golden JSON is not available.")

        payload = json.loads(path.read_text(encoding="utf-8"))
        report = compare_against_reference(payload, copy.deepcopy(payload))

        self.assertEqual(list(payload.keys()), ["gstin", "fp", "version", "hash", "b2cs", "supeco", "doc_issue"])
        self.assertEqual(payload["fp"], "042026")
        self.assertEqual([row["etin"] for row in payload["supeco"]["clttx"]], ["07AARCM9332R1CQ", "07AACCF0683K1CU", "07AAICA3918J1CV"])
        self.assertTrue(report["exact_match"])
        self.assertEqual(report["match_score"], 100.0)

    def test_excel_template_export_preserves_original_gsttool_workbook_layout(self):
        template = Path("/home/jarvis/Downloads/GSTR1_07TCRPS8655B1ZK_monthly_032026.xlsx")
        if not template.exists():
            self.skipTest("Original GSTTool March Excel workbook is not available.")

        payload = self.march_payload()
        with TemporaryDirectory() as temp_dir:
            generated_path = write_gstr1_template_excel(Path(temp_dir) / "gstr1.xlsx", payload, template)
            reference_wb = load_workbook(template)
            generated_wb = load_workbook(generated_path)
            try:
                self.assertEqual(generated_wb.sheetnames, reference_wb.sheetnames)
                for sheet_name in reference_wb.sheetnames:
                    reference = reference_wb[sheet_name]
                    generated = generated_wb[sheet_name]
                    self.assertEqual(generated.max_column, reference.max_column, sheet_name)
                    self.assertEqual(str(generated.merged_cells.ranges), str(reference.merged_cells.ranges), sheet_name)
                    self.assertEqual(generated.freeze_panes, reference.freeze_panes, sheet_name)
                    self.assertEqual(generated.auto_filter.ref, reference.auto_filter.ref, sheet_name)
                    self.assertEqual(generated.page_setup.orientation, reference.page_setup.orientation, sheet_name)
                    for col in range(1, reference.max_column + 1):
                        letter = reference.cell(1, col).column_letter
                        self.assertEqual(generated.column_dimensions[letter].width, reference.column_dimensions[letter].width, sheet_name)
                    for row in range(1, min(reference.max_row, generated.max_row) + 1):
                        self.assertEqual(generated.row_dimensions[row].height, reference.row_dimensions[row].height, sheet_name)
                    for row in range(1, min(4, reference.max_row) + 1):
                        for col in range(1, reference.max_column + 1):
                            self.assertEqual(generated.cell(row, col).style_id, reference.cell(row, col).style_id, f"{sheet_name}!{row}:{col}")

                self.assertEqual(generated_wb["b2cs"]["A5"].value, "OE")
                self.assertEqual(generated_wb["b2cs"]["B5"].value, "37-Andhra Pradesh")
                self.assertEqual(generated_wb["eco"]["B7"].value, "07AACCF0683K1CU")
                self.assertEqual(generated_wb["eco"]["F7"].value, 1.49)
                self.assertEqual(generated_wb["eco"]["G7"].value, 1.49)
                self.assertEqual(generated_wb["docs"]["A7"].value, "Invoices for outward supply")
                self.assertEqual(generated_wb["docs"]["B7"].value, "IN-5")
            finally:
                reference_wb.close()
                generated_wb.close()


if __name__ == "__main__":
    unittest.main()
